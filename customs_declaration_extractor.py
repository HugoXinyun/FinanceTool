from __future__ import annotations

import argparse
import os
import re
import tempfile
from collections import Counter
from dataclasses import dataclass, field
from datetime import date
from decimal import Decimal, InvalidOperation
from enum import Enum
from io import BytesIO
from pathlib import Path
from typing import Callable, Iterable, Sequence
from unicodedata import east_asian_width

import fitz
from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


TARGET_WIDTH = 842.0
TARGET_HEIGHT = 595.0
WEIGHT_UNITS = {"千克", "公斤", "kg", "KG"}
QUANTITY_UNITS = ("千克", "公斤", "干克", "件", "条", "个", "套", "双", "台", "米", "吨")

PRIMARY = "174A57"
SECONDARY = "2F7D6E"
SOFT = "E8F1F1"
LIGHT = "F7FAFA"
GRID = "D5DEDE"
TEXT = "203133"
WARNING = "FFF2CC"
WARNING_TEXT = "8A5700"
WHITE = "FFFFFF"


class ExtractionStatus(str, Enum):
    SUCCESS = "成功"
    REVIEW = "需复核"
    OCR_REQUIRED = "需要OCR"
    FAILED = "失败"


@dataclass(frozen=True)
class PdfWord:
    x0: float
    y0: float
    x1: float
    y1: float
    text: str
    confidence: float = 1.0


@dataclass(frozen=True)
class _OcrPageAnalysis:
    page_index: int
    words: list[PdfWord]
    text: str
    score: int
    customs_number: str
    has_title: bool
    has_item_identity: bool
    has_item_table: bool
    header_hits: int
    hs_count: int


@dataclass(frozen=True)
class Quantity:
    value: Decimal
    unit: str


@dataclass
class CustomsItem:
    item_no: int
    hs_code: str
    product_name: str
    specification: str
    quantities: list[Quantity]
    unit_price: Decimal | None
    total_price: Decimal | None
    currency: str
    origin: str
    final_destination: str
    domestic_source: str
    exemption: str
    review_notes: list[str] = field(default_factory=list)

    @property
    def weight_kg(self) -> Decimal | None:
        for quantity in self.quantities:
            if quantity.unit in WEIGHT_UNITS:
                return quantity.value
        return None

    @property
    def transaction_quantity(self) -> Quantity | None:
        for quantity in self.quantities:
            if quantity.unit not in WEIGHT_UNITS:
                return quantity
        return self.quantities[0] if self.quantities else None


@dataclass
class CustomsDeclaration:
    source_pdf: str
    pdf_pages: int
    declaration_pages: list[int]
    pre_entry_number: str
    customs_number: str
    contract_number: str
    domestic_consignor: str
    foreign_consignee: str
    exit_customs: str
    export_date: date | None
    declaration_date: date | None
    transport_mode: str
    transport_tool: str
    bill_of_lading_no: str
    supervision_mode: str
    levy_nature: str
    trade_country: str
    destination_country: str
    destination_port: str
    exit_port: str
    package_type: str
    packages: int | None
    gross_weight_kg: Decimal | None
    net_weight_kg: Decimal | None
    terms: str
    freight: Decimal | None
    insurance: Decimal | None
    miscellaneous: Decimal | None
    remarks: str
    items: list[CustomsItem]
    review_notes: list[str] = field(default_factory=list)

    @property
    def total_item_weight_kg(self) -> Decimal:
        return sum((item.weight_kg or Decimal("0") for item in self.items), Decimal("0"))

    @property
    def declared_totals(self) -> dict[str, Decimal]:
        totals: dict[str, Decimal] = {}
        for item in self.items:
            if item.total_price is None:
                continue
            currency = item.currency or "未注明"
            totals[currency] = totals.get(currency, Decimal("0")) + item.total_price
        return totals


@dataclass
class ExtractionResult:
    source_pdf: str
    status: ExtractionStatus
    declaration: CustomsDeclaration | None = None
    messages: list[str] = field(default_factory=list)


@dataclass
class ExtractionBatch:
    results: list[ExtractionResult]

    @property
    def declarations(self) -> list[CustomsDeclaration]:
        return [result.declaration for result in self.results if result.declaration is not None]

    @property
    def items(self) -> list[CustomsItem]:
        return [item for declaration in self.declarations for item in declaration.items]

    @property
    def review_count(self) -> int:
        return sum(result.status == ExtractionStatus.REVIEW for result in self.results)

    @property
    def failure_count(self) -> int:
        return sum(result.status in {ExtractionStatus.FAILED, ExtractionStatus.OCR_REQUIRED} for result in self.results)


ProgressCallback = Callable[[int, int, ExtractionResult], None]


FIELD_ALIASES: dict[str, tuple[str, ...]] = {
    "domestic_consignor": ("境内发货人",),
    "foreign_consignee": ("境外收货人", "墙外收货人"),
    "exit_customs": ("出境关别",),
    "export_date": ("出口日期",),
    "declaration_date": ("申报日期",),
    "transport_mode": ("运输方式",),
    "transport_tool": ("运输工具名称及航次号", "运输工具名称"),
    "bill_of_lading_no": ("提运单号",),
    "supervision_mode": ("监管方式",),
    "levy_nature": ("征免性质",),
    "trade_country": ("贸易国(地区)", "贸易国（地区）", "贸易国"),
    "destination_country": ("运抵国(地区)", "运抵国（地区）", "运抵国"),
    "destination_port": ("指运港",),
    "exit_port": ("离境口岸",),
    "contract_number": ("合同协议号",),
    "package_type": ("包装种类",),
    "packages": ("件数",),
    "gross_weight_kg": ("毛重(千克)", "毛重（千克）", "毛重(干克)", "毛重"),
    "net_weight_kg": ("净重(千克)", "净重（千克）", "净重(干克)", "净重"),
    "terms": ("成交方式",),
    "freight": ("运费",),
    "insurance": ("保费",),
    "miscellaneous": ("杂费",),
}

ALL_LABELS = tuple(alias for aliases in FIELD_ALIASES.values() for alias in aliases) + (
    "生产销售单位",
    "备案号",
    "许可证号",
    "随附单证及编号",
    "标记唛码及备注",
    "标记喷码及备注",
)


def _normalized_text(value: str) -> str:
    return (
        value.replace("（", "(")
        .replace("）", ")")
        .replace("：", ":")
        .replace("｜", "|")
        .replace("丨", "|")
        .replace(" ", "")
        .strip()
    )


def _decimal(value: str | int | float | Decimal | None) -> Decimal | None:
    if value is None or value == "":
        return None
    if isinstance(value, Decimal):
        return value
    cleaned = str(value).replace(",", "").strip()
    try:
        return Decimal(cleaned)
    except InvalidOperation:
        return None


def _date(value: str) -> date | None:
    match = re.search(r"(20\d{2})[./年-]?(\d{2})[./月-]?(\d{2})", value)
    if not match:
        return None
    try:
        return date(int(match.group(1)), int(match.group(2)), int(match.group(3)))
    except ValueError:
        return None


def _join_tokens(tokens: Sequence[str]) -> str:
    result = ""
    for token in tokens:
        token = token.strip()
        if not token:
            continue
        if result and result[-1].isascii() and result[-1].isalnum() and token[0].isascii() and token[0].isalnum():
            result += " "
        result += token
    return result.strip()


def _group_lines(words: Sequence[PdfWord], tolerance: float = 2.8) -> list[tuple[float, list[PdfWord]]]:
    lines: list[tuple[float, list[PdfWord]]] = []
    for word in sorted(words, key=lambda item: (item.y0, item.x0)):
        if not lines or abs(lines[-1][0] - word.y0) > tolerance:
            lines.append((word.y0, [word]))
        else:
            lines[-1][1].append(word)
    for _, line_words in lines:
        line_words.sort(key=lambda item: item.x0)
    return lines


def _region_text(words: Sequence[PdfWord], x0: float, x1: float, y0: float, y1: float) -> str:
    selected = [word for word in words if x0 <= word.x0 < x1 and y0 <= word.y0 < y1]
    return " ".join(_join_tokens([word.text for word in line]) for _, line in _group_lines(selected)).strip()


def _visual_words(page: fitz.Page) -> list[PdfWord]:
    matrix = page.rotation_matrix
    width = float(page.rect.width)
    height = float(page.rect.height)
    scale_x = TARGET_WIDTH / width
    scale_y = TARGET_HEIGHT / height
    result: list[PdfWord] = []
    for raw_word in page.get_text("words", sort=False):
        rect = fitz.Rect(raw_word[:4]) * matrix
        result.append(
            PdfWord(
                x0=min(rect.x0, rect.x1) * scale_x,
                y0=min(rect.y0, rect.y1) * scale_y,
                x1=max(rect.x0, rect.x1) * scale_x,
                y1=max(rect.y0, rect.y1) * scale_y,
                text=str(raw_word[4]).strip(),
            )
        )
    return [word for word in result if word.text]


def _declaration_page_score(words: Sequence[PdfWord]) -> int:
    text = _normalized_text("".join(word.text for word in words))
    score = 0
    if "中华人民共和国海关出口" in text:
        score += 4
    if "报关单" in text:
        score += 2
    if "项号" in text or ("商品编号" in text and "数量及单位" in text):
        score += 2
    if re.search(r"(?<!\d)\d{18}(?!\d)", text):
        score += 1
    if re.search(r"(?<!\d)\d{10}(?!\d)", text):
        score += 1
    return score


def _find_label(words: Sequence[PdfWord], aliases: Sequence[str]) -> PdfWord | None:
    normalized_aliases = tuple(_normalized_text(alias) for alias in aliases)
    matches = [
        word
        for word in words
        if any(alias in _normalized_text(word.text) for alias in normalized_aliases)
    ]
    return min(matches, key=lambda word: (word.y0, word.x0)) if matches else None


def _is_known_label(text: str) -> bool:
    normalized = _normalized_text(text)
    return any(_normalized_text(alias) in normalized for alias in ALL_LABELS)


def _cell_right(words: Sequence[PdfWord], label: PdfWord) -> float:
    candidates = [
        word.x0
        for word in words
        if word.x0 > label.x0 + 15
        and abs(word.y0 - label.y0) <= 5
        and _is_known_label(word.text)
    ]
    return min(candidates) - 3 if candidates else TARGET_WIDTH


def _cell_value_below(words: Sequence[PdfWord], label: PdfWord | None) -> str:
    if label is None:
        return ""
    right = _cell_right(words, label)
    next_label_y = min(
        (
            word.y0
            for word in words
            if word.y0 > label.y0 + 5
            and label.x0 - 3 <= word.x0 < right
            and _is_known_label(word.text)
        ),
        default=label.y0 + 32,
    )
    end_y = min(label.y0 + 31, next_label_y - 5)
    selected = [
        word
        for word in words
        if label.x0 - 3 <= word.x0 < right
        and label.y0 + 5 < word.y0 <= end_y
        and not _is_known_label(word.text)
    ]
    lines = _group_lines(selected)
    return " ".join(_join_tokens([word.text for word in line]) for _, line in lines).strip()


def _code_near_label(words: Sequence[PdfWord], label: PdfWord | None) -> str:
    if label is None:
        return ""
    right = _cell_right(words, label)
    candidates = [label.text] + [
        word.text
        for word in words
        if label.x0 - 3 <= word.x0 < right and label.y0 - 7 <= word.y0 <= label.y0 + 5
    ]
    for candidate in candidates:
        match = re.search(r"[（(]([A-Za-z0-9]+)[）)]", candidate)
        if match:
            return match.group(1)
    return ""


def _labelled_value(words: Sequence[PdfWord], field_name: str) -> tuple[str, str]:
    label = _find_label(words, FIELD_ALIASES[field_name])
    return _cell_value_below(words, label), _code_near_label(words, label)


def _format_code_name(code: str, name: str) -> str:
    code = code.strip()
    name = name.strip()
    if code and name:
        return f"({code}) {name}"
    return name or code


def _extract_number_after(text: str, label_pattern: str, length: str) -> str:
    compact = _normalized_text(text)
    match = re.search(rf"{label_pattern}[:]?([A-Z]?\d{{{length}}})", compact, flags=re.IGNORECASE)
    return match.group(1) if match else ""


def _most_common_customs_number(text: str) -> str:
    numbers = re.findall(r"(?<!\d)(\d{18})(?!\d)", _normalized_text(text))
    return Counter(numbers).most_common(1)[0][0] if numbers else ""


def _parse_quantity_lines(words: Sequence[PdfWord]) -> list[Quantity]:
    quantities: list[Quantity] = []
    selected = [word for word in words if 315 <= word.x0 < 430]
    pattern = re.compile(rf"(\d+(?:\.\d+)?)\s*({'|'.join(QUANTITY_UNITS)})", re.IGNORECASE)
    for _, line in _group_lines(selected):
        text = _join_tokens([word.text for word in line]).replace("干克", "千克")
        match = pattern.search(text)
        if match:
            value = _decimal(match.group(1))
            if value is not None:
                quantities.append(Quantity(value=value, unit=match.group(2).replace("干克", "千克")))
    return quantities


def _parse_price_lines(words: Sequence[PdfWord]) -> list[Decimal]:
    prices: list[Decimal] = []
    for _, line in _group_lines([word for word in words if 425 <= word.x0 < 530]):
        text = _join_tokens([word.text for word in line]).replace(",", "")
        for match in re.finditer(r"(?<!\d)(\d+(?:\.\d+)?)(?!\d)", text):
            value = _decimal(match.group(1))
            if value is not None:
                prices.append(value)
                break
    return prices


def _split_product_text(words: Sequence[PdfWord], hs_code: str) -> tuple[str, str]:
    prefixes: list[str] = []
    for word in words:
        if hs_code and hs_code in word.text:
            remainder = word.text.split(hs_code, 1)[1].strip()
            if remainder:
                prefixes.append(remainder)
    product_words = [word for word in words if 100 <= word.x0 < 330]
    lines = [_join_tokens([word.text for word in line]) for _, line in _group_lines(product_words)]
    text = "".join(prefixes + lines).strip()
    marker = re.search(r"3[|｜丨][0OoＯ]", text)
    if marker:
        product_name = text[: marker.start()].strip(" |")
        specification = text[marker.start() :]
    else:
        product_name = text if len(text) <= 40 else ""
        specification = "" if product_name else text
    specification = (
        specification.replace("｜", "|")
        .replace("丨", "|")
        .replace("3|O", "3|0")
        .replace("3|o", "3|0")
    )
    return product_name, specification


def _parse_item_block(item_no: int, words: Sequence[PdfWord]) -> CustomsItem:
    block_text = " ".join(word.text for word in words)
    hs_match = re.search(r"(?<!\d)(\d{10})(?!\d)", block_text)
    if not hs_match:
        hs_match = re.search(r"(\d{10})", block_text)
    hs_code = hs_match.group(1) if hs_match else ""
    product_name, specification = _split_product_text(words, hs_code)
    quantities = _parse_quantity_lines(words)
    prices = _parse_price_lines(words)
    unit_price: Decimal | None = None
    total_price: Decimal | None = None
    notes: list[str] = []
    if len(prices) >= 2:
        unit_price, total_price = prices[0], prices[1]
    elif len(prices) == 1:
        total_price = prices[0]
        transaction = next((quantity for quantity in quantities if quantity.unit not in WEIGHT_UNITS), None)
        if transaction and transaction.value:
            unit_price = total_price / transaction.value
            notes.append("单价由总价除以数量推算")
        else:
            notes.append("价格字段不完整")

    currency = "美元" if "美元" in block_text or "USD" in block_text.upper() else ""
    origin = _region_text(words, 525, 620, 0, TARGET_HEIGHT)
    destination = _region_text(words, 610, 700, 0, TARGET_HEIGHT)
    domestic_source = _region_text(words, 680, 765, 0, TARGET_HEIGHT)
    exemption = "照章征税 (1)" if "照章征税" in block_text else _region_text(words, 750, 830, 0, TARGET_HEIGHT)
    domestic_source = domestic_source.replace("照章征税", "").strip()

    if not hs_code:
        notes.append("未识别商品编号")
    if not product_name:
        notes.append("未可靠识别商品名称")
    if not quantities:
        notes.append("未识别数量及单位")
    if "?" in specification or "�" in specification:
        notes.append("规格文本包含OCR不确定字符")
    if specification.endswith("，") or specification.endswith("不是套") or specification.endswith("|不"):
        notes.append("规格在源PDF中疑似截断")

    return CustomsItem(
        item_no=item_no,
        hs_code=hs_code,
        product_name=product_name,
        specification=specification,
        quantities=quantities,
        unit_price=unit_price,
        total_price=total_price,
        currency=currency,
        origin=origin,
        final_destination=destination,
        domestic_source=domestic_source,
        exemption=exemption,
        review_notes=notes,
    )


def _parse_page_items(words: Sequence[PdfWord]) -> list[CustomsItem]:
    item_header = _find_label(words, ("项号",))
    if item_header is None:
        return []
    header_y = item_header.y0
    anchors: list[tuple[int, float]] = []
    for word in words:
        match = re.fullmatch(r"\s*(\d{1,2})\s*", word.text)
        if not match or word.x0 > 90 or word.y0 <= header_y + 3:
            continue
        nearby = " ".join(
            candidate.text
            for candidate in words
            if abs(candidate.y0 - word.y0) <= 8 and candidate.x0 < 300
        )
        if re.search(r"\d{10}", nearby):
            anchors.append((int(match.group(1)), word.y0))
    anchors = sorted(set(anchors), key=lambda value: value[1])

    footer_markers = (
        "特殊关系",
        "特莱关系",
        "支付特许权",
        "公式定价",
        "暂定价格",
        "报关人员",
        "申报单位",
        "自报自缴",
    )
    items: list[CustomsItem] = []
    for index, (item_no, start_y) in enumerate(anchors):
        next_y = anchors[index + 1][1] if index + 1 < len(anchors) else TARGET_HEIGHT
        footer_y = min(
            (
                word.y0
                for word in words
                if word.y0 > start_y + 5 and any(marker in word.text for marker in footer_markers)
            ),
            default=TARGET_HEIGHT,
        )
        end_y = min(next_y, footer_y)
        block = [word for word in words if start_y - 3 <= word.y0 < end_y - 1]
        items.append(_parse_item_block(item_no, block))
    return items


def _is_scan_like(page: fitz.Page) -> bool:
    words = page.get_text("words", sort=False)
    text = " ".join(str(word[4]) for word in words)
    suspicious_markers = ("?", "商品编玛", "干克", "墙外收货人", "惨织", "申摇", "特莱关系")
    text_is_suspicious = len(words) < 120 or any(marker in text for marker in suspicious_markers)
    if not text_is_suspicious:
        return False
    page_area = float(page.rect.width * page.rect.height)
    for image in page.get_images(full=True):
        try:
            rects = page.get_image_rects(image[0])
        except Exception:
            continue
        if any(float(rect.width * rect.height) / page_area > 0.85 for rect in rects):
            return True
    return False


def _remarks(words: Sequence[PdfWord]) -> str:
    label = _find_label(words, ("标记唛码及备注", "标记喷码及备注", "备注"))
    if label is None:
        return ""
    item_header = _find_label(words, ("项号",))
    end_y = item_header.y0 - 2 if item_header else min(TARGET_HEIGHT, label.y0 + 55)
    text = _region_text(words, 25, TARGET_WIDTH - 20, label.y0, end_y)
    text = re.sub(r"^标记[唛喷]码及备注", "", text).strip()
    text = re.sub(r"^备注[:：]?", "", text).strip()
    return text


def _extract_header(
    source_path: Path,
    pdf_pages: int,
    declaration_pages: list[int],
    words: Sequence[PdfWord],
    all_text: str,
    items: list[CustomsItem],
) -> CustomsDeclaration:
    compact = _normalized_text(all_text)
    pre_entry = _extract_number_after(all_text, "预录入编号", "15,25")
    customs_number = _extract_number_after(all_text, "海关编号", "18") or _most_common_customs_number(all_text)
    contract_value, _ = _labelled_value(words, "contract_number")
    contract_match = re.search(r"[A-Z]\d{7}-\d{3}", all_text, flags=re.IGNORECASE)
    filename_contract = re.search(r"[A-Z]\d{7}-\d{3}", source_path.stem, flags=re.IGNORECASE)
    contract_number = (
        contract_match.group(0)
        if contract_match
        else contract_value
        if re.fullmatch(r"[A-Z]\d{7}-\d{3}", contract_value, flags=re.IGNORECASE)
        else filename_contract.group(0)
        if filename_contract
        else source_path.stem
    )

    domestic_consignor, _ = _labelled_value(words, "domestic_consignor")
    foreign_consignee, _ = _labelled_value(words, "foreign_consignee")
    exit_customs_name, exit_customs_code = _labelled_value(words, "exit_customs")
    export_date_text, _ = _labelled_value(words, "export_date")
    declaration_date_text, _ = _labelled_value(words, "declaration_date")
    transport_name, transport_code = _labelled_value(words, "transport_mode")
    transport_tool, _ = _labelled_value(words, "transport_tool")
    bill_of_lading_no, _ = _labelled_value(words, "bill_of_lading_no")
    supervision_name, supervision_code = _labelled_value(words, "supervision_mode")
    levy_name, levy_code = _labelled_value(words, "levy_nature")
    trade_name, trade_code = _labelled_value(words, "trade_country")
    destination_name, destination_code = _labelled_value(words, "destination_country")
    port_name, port_code = _labelled_value(words, "destination_port")
    exit_port_name, exit_port_code = _labelled_value(words, "exit_port")
    package_name, package_code = _labelled_value(words, "package_type")
    packages_text, _ = _labelled_value(words, "packages")
    gross_text, _ = _labelled_value(words, "gross_weight_kg")
    net_text, _ = _labelled_value(words, "net_weight_kg")
    terms_name, terms_code = _labelled_value(words, "terms")
    freight_text, _ = _labelled_value(words, "freight")
    insurance_text, _ = _labelled_value(words, "insurance")
    misc_text, _ = _labelled_value(words, "miscellaneous")

    declaration_date = _date(declaration_date_text)
    if declaration_date is None:
        date_candidates = re.findall(r"20\d{6}", compact)
        declaration_date = _date(date_candidates[0]) if date_candidates else None
    packages_value = _decimal(re.search(r"\d+", packages_text).group(0)) if re.search(r"\d+", packages_text) else None

    return CustomsDeclaration(
        source_pdf=source_path.name,
        pdf_pages=pdf_pages,
        declaration_pages=declaration_pages,
        pre_entry_number=pre_entry or customs_number,
        customs_number=customs_number,
        contract_number=contract_number,
        domestic_consignor=domestic_consignor,
        foreign_consignee=foreign_consignee,
        exit_customs=_format_code_name(exit_customs_code, exit_customs_name),
        export_date=_date(export_date_text),
        declaration_date=declaration_date,
        transport_mode=_format_code_name(transport_code, transport_name),
        transport_tool=transport_tool,
        bill_of_lading_no=bill_of_lading_no,
        supervision_mode=_format_code_name(supervision_code, supervision_name),
        levy_nature=_format_code_name(levy_code, levy_name),
        trade_country=_format_code_name(trade_code, trade_name),
        destination_country=_format_code_name(destination_code, destination_name),
        destination_port=_format_code_name(port_code, port_name),
        exit_port=_format_code_name(exit_port_code, exit_port_name),
        package_type=_format_code_name(package_code, package_name),
        packages=int(packages_value) if packages_value is not None else None,
        gross_weight_kg=_decimal(re.search(r"\d+(?:\.\d+)?", gross_text).group(0)) if re.search(r"\d+(?:\.\d+)?", gross_text) else None,
        net_weight_kg=_decimal(re.search(r"\d+(?:\.\d+)?", net_text).group(0)) if re.search(r"\d+(?:\.\d+)?", net_text) else None,
        terms=_format_code_name(terms_code, terms_name),
        freight=_decimal(re.search(r"\d+(?:\.\d+)?", freight_text).group(0)) if re.search(r"\d+(?:\.\d+)?", freight_text) else None,
        insurance=_decimal(re.search(r"\d+(?:\.\d+)?", insurance_text).group(0)) if re.search(r"\d+(?:\.\d+)?", insurance_text) else None,
        miscellaneous=_decimal(re.search(r"\d+(?:\.\d+)?", misc_text).group(0)) if re.search(r"\d+(?:\.\d+)?", misc_text) else None,
        remarks=_remarks(words),
        items=items,
    )


def _validate_declaration(declaration: CustomsDeclaration) -> list[str]:
    notes: list[str] = []
    if not re.fullmatch(r"\d{18}", declaration.customs_number):
        notes.append("海关编号不是18位数字或未识别")
    if declaration.declaration_date is None:
        notes.append("未识别申报日期")
    if not declaration.domestic_consignor:
        notes.append("未识别境内发货人")
    if not declaration.foreign_consignee:
        notes.append("未识别境外收货人")
    if declaration.net_weight_kg is None:
        notes.append("未识别报关单净重")
    if not declaration.items:
        notes.append("未识别商品明细")
    else:
        item_numbers = [item.item_no for item in declaration.items]
        expected = list(range(1, max(item_numbers) + 1))
        if item_numbers != expected:
            notes.append(f"商品项号不连续：{item_numbers}")

    if declaration.net_weight_kg is not None:
        difference = declaration.total_item_weight_kg - declaration.net_weight_kg
        if abs(difference) > Decimal("0.05"):
            notes.append(f"商品重量合计与净重相差{difference}千克")

    currencies = {item.currency for item in declaration.items if item.currency}
    if len(currencies) > 1:
        notes.append("同一报关单包含多种币制，金额需分币制核对")

    for item in declaration.items:
        if not re.fullmatch(r"\d{10}", item.hs_code):
            notes.append(f"第{item.item_no}项商品编号无效")
        if len(item.quantities) < 2:
            notes.append(f"第{item.item_no}项数量及单位不完整")
        if item.weight_kg is None:
            notes.append(f"第{item.item_no}项未识别千克重量")
        if item.unit_price is None:
            notes.append(f"第{item.item_no}项未识别单价")
        if item.total_price is None:
            notes.append(f"第{item.item_no}项未识别总价")
        if not item.currency:
            notes.append(f"第{item.item_no}项未识别币制")
        transaction = item.transaction_quantity
        if transaction and item.unit_price is not None and item.total_price is not None:
            calculated = transaction.value * item.unit_price
            unit_dp = max(0, -item.unit_price.as_tuple().exponent)
            total_dp = max(0, -item.total_price.as_tuple().exponent)
            tolerance = (
                Decimal("0.5") * transaction.value * (Decimal(10) ** -unit_dp)
                + Decimal("0.5") * (Decimal(10) ** -total_dp)
                + Decimal("0.01")
            )
            if abs(calculated - item.total_price) > tolerance:
                notes.append(f"第{item.item_no}项数量乘单价与总价不符")
        notes.extend(f"第{item.item_no}项：{note}" for note in item.review_notes)
    return list(dict.fromkeys(notes))


def _extract_largest_page_image(document: fitz.Document, page: fitz.Page):
    from PIL import Image

    candidates = []
    for image_info in page.get_images(full=True):
        try:
            image_data = document.extract_image(image_info[0])
        except Exception:
            continue
        candidates.append((image_data.get("width", 0) * image_data.get("height", 0), image_data))
    if candidates:
        image_data = max(candidates, key=lambda item: item[0])[1]
        image = Image.open(BytesIO(image_data["image"])).convert("RGB")
    else:
        pixmap = page.get_pixmap(matrix=fitz.Matrix(3, 3), alpha=False)
        image = Image.frombytes("RGB", (pixmap.width, pixmap.height), pixmap.samples)
    if image.height > image.width:
        image = image.rotate(90, expand=True)
    return image


class CustomsDeclarationExtractor:
    def __init__(self, enable_ocr: bool = True):
        self.enable_ocr = enable_ocr
        self._ocr_reader = None

    def extract_files(
        self,
        paths: Iterable[str | os.PathLike[str]],
        progress_callback: ProgressCallback | None = None,
    ) -> ExtractionBatch:
        files = self.expand_inputs(paths)
        results: list[ExtractionResult] = []
        for index, path in enumerate(files, start=1):
            result = self.extract_pdf(path)
            results.append(result)
            if progress_callback:
                progress_callback(index, len(files), result)
        return ExtractionBatch(results)

    @staticmethod
    def expand_inputs(paths: Iterable[str | os.PathLike[str]]) -> list[Path]:
        files: list[Path] = []
        seen: set[str] = set()
        for raw_path in paths:
            path = Path(raw_path).expanduser()
            candidates = sorted(path.glob("*.pdf")) if path.is_dir() else [path]
            for candidate in candidates:
                if candidate.suffix.lower() != ".pdf":
                    continue
                normalized = os.path.normcase(str(candidate.resolve()))
                if normalized not in seen:
                    seen.add(normalized)
                    files.append(candidate.resolve())
        return files

    def extract_pdf(self, path: str | os.PathLike[str]) -> ExtractionResult:
        source_path = Path(path)
        if not source_path.exists():
            return ExtractionResult(source_path.name, ExtractionStatus.FAILED, messages=["文件不存在"])
        if source_path.suffix.lower() != ".pdf":
            return ExtractionResult(source_path.name, ExtractionStatus.FAILED, messages=["不是PDF文件"])

        try:
            document = fitz.open(source_path)
        except Exception as exc:
            return ExtractionResult(source_path.name, ExtractionStatus.FAILED, messages=[f"无法打开PDF：{exc}"])

        try:
            if document.needs_pass:
                return ExtractionResult(source_path.name, ExtractionStatus.FAILED, messages=["PDF已加密"])
            page_words = [_visual_words(page) for page in document]
            declaration_indexes = [
                index for index, words in enumerate(page_words) if _declaration_page_score(words) >= 6
            ]
            if not declaration_indexes:
                if not self.enable_ocr:
                    return ExtractionResult(
                        source_path.name,
                        ExtractionStatus.OCR_REQUIRED,
                        messages=["未找到可靠文本层，需要OCR或人工复核"],
                    )
                return self._extract_with_ocr(document, source_path)

            first_index = declaration_indexes[0]
            first_words = page_words[first_index]
            all_words = [word for index in declaration_indexes for word in page_words[index]]
            all_text = "\n".join(word.text for word in all_words)
            items = [item for index in declaration_indexes for item in _parse_page_items(page_words[index])]
            items.sort(key=lambda item: item.item_no)
            declaration = _extract_header(
                source_path=source_path,
                pdf_pages=len(document),
                declaration_pages=[index + 1 for index in declaration_indexes],
                words=first_words,
                all_text=all_text,
                items=items,
            )
            if _is_scan_like(document[first_index]):
                declaration.review_notes.append("报关单为整页图像并带隐藏文本层，规格文本需复核")
            declaration.review_notes.extend(_validate_declaration(declaration))
            declaration.review_notes = list(dict.fromkeys(declaration.review_notes))
            status = ExtractionStatus.REVIEW if declaration.review_notes else ExtractionStatus.SUCCESS
            return ExtractionResult(source_path.name, status, declaration, declaration.review_notes.copy())
        except Exception as exc:
            return ExtractionResult(source_path.name, ExtractionStatus.FAILED, messages=[f"解析失败：{exc}"])
        finally:
            document.close()

    def _get_ocr_reader(self):
        if self._ocr_reader is not None:
            return self._ocr_reader
        try:
            import easyocr
        except ImportError as exc:
            raise RuntimeError("未安装EasyOCR，无法识别扫描版报关单") from exc
        self._ocr_reader = easyocr.Reader(
            ["ch_sim", "en"],
            gpu=False,
            verbose=False,
            download_enabled=False,
        )
        return self._ocr_reader

    def _ocr_words(self, image) -> list[PdfWord]:
        import numpy as np
        from PIL import ImageEnhance, ImageOps

        max_width = 1800
        if image.width > max_width:
            ratio = max_width / image.width
            image = image.resize((max_width, int(image.height * ratio)))
        prepared = ImageEnhance.Contrast(ImageOps.grayscale(image)).enhance(1.4)
        result = self._get_ocr_reader().readtext(
            np.asarray(prepared),
            detail=1,
            paragraph=False,
            canvas_size=2200,
            mag_ratio=1.0,
        )
        words: list[PdfWord] = []
        for box, text, confidence in result:
            xs = [point[0] for point in box]
            ys = [point[1] for point in box]
            words.append(
                PdfWord(
                    min(xs) / prepared.width * TARGET_WIDTH,
                    min(ys) / prepared.height * TARGET_HEIGHT,
                    max(xs) / prepared.width * TARGET_WIDTH,
                    max(ys) / prepared.height * TARGET_HEIGHT,
                    text.strip(),
                    float(confidence),
                )
            )
        return words

    def _ocr_crop(
        self,
        image,
        bounds: tuple[float, float, float, float],
        allowlist: str | None = None,
        enlarge: float = 1.0,
        threshold: int | None = None,
    ) -> list[PdfWord]:
        import numpy as np
        from PIL import ImageEnhance, ImageOps

        x0, y0, x1, y1 = bounds
        scale_x = image.width / TARGET_WIDTH
        scale_y = image.height / TARGET_HEIGHT
        crop = image.crop((int(x0 * scale_x), int(y0 * scale_y), int(x1 * scale_x), int(y1 * scale_y)))
        prepared = ImageOps.autocontrast(ImageOps.grayscale(crop))
        prepared = ImageEnhance.Contrast(prepared).enhance(2.0)
        if threshold is not None:
            prepared = prepared.point(lambda pixel: 0 if pixel < threshold else 255)
        if enlarge != 1.0:
            prepared = prepared.resize((int(prepared.width * enlarge), int(prepared.height * enlarge)))
        result = self._get_ocr_reader().readtext(
            np.asarray(prepared),
            detail=1,
            paragraph=False,
            canvas_size=3000,
            mag_ratio=1.0,
            allowlist=allowlist,
        )
        words: list[PdfWord] = []
        for box, text, confidence in result:
            xs = [point[0] / enlarge for point in box]
            ys = [point[1] / enlarge for point in box]
            words.append(
                PdfWord(
                    x0 + min(xs) / scale_x,
                    y0 + min(ys) / scale_y,
                    x0 + max(xs) / scale_x,
                    y0 + max(ys) / scale_y,
                    text.strip(),
                    float(confidence),
                )
            )
        return words

    @staticmethod
    def _ocr_hs_entries(words: Sequence[PdfWord]) -> list[tuple[float, str, float]]:
        entries: list[tuple[float, str, float]] = []
        for word in sorted(words, key=lambda value: value.y0):
            sequences = re.findall(r"\d{10,}", word.text)
            if not sequences:
                continue
            code = sequences[0][:10]
            center_y = (word.y0 + word.y1) / 2
            if entries and abs(entries[-1][0] - center_y) < 3:
                if word.confidence > entries[-1][2]:
                    entries[-1] = (center_y, code, word.confidence)
                continue
            entries.append((center_y, code, word.confidence))
        return entries

    @staticmethod
    def _ocr_quantities(words: Sequence[PdfWord]) -> list[tuple[float, Quantity, float]]:
        pattern = re.compile(rf"(\d+(?:\.\d+)?)\s*({'|'.join(QUANTITY_UNITS)})", re.IGNORECASE)
        result: list[tuple[float, Quantity, float]] = []
        for word in sorted(words, key=lambda value: value.y0):
            text = word.text.replace("干克", "千克")
            match = pattern.search(text)
            value = _decimal(match.group(1)) if match else None
            if match and value is not None:
                result.append(((word.y0 + word.y1) / 2, Quantity(value, match.group(2).replace("干克", "千克")), word.confidence))
        return result

    @staticmethod
    def _ocr_prices(words: Sequence[PdfWord]) -> list[tuple[float, Decimal, float]]:
        result: list[tuple[float, Decimal, float]] = []
        for word in sorted(words, key=lambda value: value.y0):
            match = re.search(r"(?<!\d)(\d+(?:\.\d{1,4})?)(?!\d)", word.text)
            value = _decimal(match.group(1)) if match else None
            if value is not None:
                result.append(((word.y0 + word.y1) / 2, value, word.confidence))
        return result

    @staticmethod
    def _ocr_row_bounds(
        code_entries: Sequence[tuple[float, str, float]],
        table_y: float,
        bottom_y: float,
    ) -> list[tuple[float, float]]:
        ordered = sorted(code_entries, key=lambda entry: entry[0])
        bounds: list[tuple[float, float]] = []
        for index, (code_y, _, _) in enumerate(ordered):
            start_y = max(table_y, code_y - 8)
            end_y = bottom_y if index + 1 == len(ordered) else ordered[index + 1][0] - 8
            bounds.append((start_y, max(start_y + 1, end_y)))
        return bounds

    @staticmethod
    def _ocr_entries_in_row(entries: Sequence[tuple], start_y: float, end_y: float) -> list[tuple]:
        return [entry for entry in entries if start_y <= entry[0] < end_y]

    @staticmethod
    def _analyze_ocr_page(page_index: int, words: list[PdfWord]) -> _OcrPageAnalysis:
        text = "\n".join(word.text for word in words)
        compact = _normalized_text("".join(word.text for word in words))
        customs_number = _extract_number_after(text, "海关编号", "18") or _most_common_customs_number(text)
        has_title = (
            ("海关出口" in compact and "报关单" in compact)
            or "出口货物报关单" in compact
        )
        hs_count = len(re.findall(r"(?<!\d)\d{10}(?!\d)", compact))
        item_markers = sum(marker in compact for marker in ("项号", "商品编号", "数量及单位", "单价", "总价"))
        has_item_identity = "项号" in compact or "商品编号" in compact
        has_item_table = has_item_identity and (item_markers >= 2 or hs_count > 0)
        header_markers = (
            "预录入编号", "海关编号", "境内发货人", "境外收货人", "申报日期",
            "运输方式", "监管方式", "合同协议号", "件数", "净重",
        )
        header_hits = sum(marker in compact for marker in header_markers)
        score = _declaration_page_score(words)
        if has_title:
            score = max(score, 6)
        if customs_number and has_item_table:
            score = max(score, 6)
        elif header_hits >= 4 and has_item_table:
            score = max(score, 5)
        return _OcrPageAnalysis(
            page_index=page_index,
            words=words,
            text=text,
            score=score,
            customs_number=customs_number,
            has_title=has_title,
            has_item_identity=has_item_identity,
            has_item_table=has_item_table,
            header_hits=header_hits,
            hs_count=hs_count,
        )

    @staticmethod
    def _select_ocr_pages(
        analyses: Sequence[_OcrPageAnalysis],
    ) -> tuple[list[_OcrPageAnalysis], list[str]]:
        warnings: list[str] = []
        reliable = [
            page
            for page in analyses
            if page.has_title
            or (bool(page.customs_number) and page.has_item_table and page.header_hits >= 2)
            or (page.header_hits >= 4 and page.has_item_table)
        ]
        if reliable:
            anchor = min(reliable, key=lambda page: page.page_index)
        else:
            fallback = [page for page in analyses if page.has_item_table and page.hs_count > 0]
            if not fallback:
                return [], warnings
            anchor = max(fallback, key=lambda page: (page.score, page.header_hits, -page.page_index))
            warnings.append(f"第{anchor.page_index + 1}页未可靠识别报关单标题，按商品表格特征提取")

        anchor_number = anchor.customs_number
        by_index = {page.page_index: page for page in analyses}

        def is_compatible(page: _OcrPageAnalysis) -> bool:
            if anchor_number and page.customs_number and page.customs_number != anchor_number:
                return False
            return page.has_title or page.has_item_table

        selected_indexes = {anchor.page_index}
        for direction in (-1, 1):
            page_index = anchor.page_index + direction
            while page_index in by_index and is_compatible(by_index[page_index]):
                selected_indexes.add(page_index)
                page_index += direction

        selected = [by_index[index] for index in sorted(selected_indexes)]
        other_reliable = [page.page_index + 1 for page in reliable if page.page_index not in selected_indexes]
        if other_reliable:
            warnings.append(f"检测到其他可能的报关单页但未合并：{other_reliable}")
        return selected, warnings

    @staticmethod
    def _detected_currency(text: str) -> str:
        currency_patterns = (
            ("美元", ("美元", "USD")),
            ("欧元", ("欧元", "EUR")),
            ("英镑", ("英镑", "GBP")),
            ("港币", ("港币", "HKD")),
            ("日元", ("日元", "JPY")),
            ("人民币", ("人民币", "CNY")),
        )
        matches = [name for name, aliases in currency_patterns if any(alias in text.upper() for alias in aliases)]
        return matches[0] if len(matches) == 1 else ""

    @staticmethod
    def _ocr_product_text(words: Sequence[PdfWord], start_y: float, end_y: float) -> tuple[str, str]:
        selected = [word for word in words if start_y <= word.y0 < end_y and not re.fullmatch(r"\d+(?:\.\d+)?", word.text)]
        text = "".join(_join_tokens([word.text for word in line]) for _, line in _group_lines(selected)).strip()
        garment_terms = ("短裤", "长裤", "半身裙", "连衣裙", "开襟衫", "套头衫", "吊带衫")
        term = next((candidate for candidate in garment_terms if candidate in text), "")
        product_name = f"女士梭织{term}" if term else ""
        marker = re.search(r"3(?:[|/1I]|10)[0Oo]?", text)
        specification = text[marker.start() :] if marker else text
        specification = re.sub(r"^3(?:[/1I]|10)[0Oo]?", "3|0|", specification)
        specification = specification.replace("310|", "3|0|").replace("3/0", "3|0|")
        specification = re.sub(r"^\d{1,12}", "", specification).strip()
        return product_name, specification

    def _extract_ocr_items(
        self,
        image,
        full_words: Sequence[PdfWord],
        default_table_y: float = 270.0,
    ) -> tuple[list[CustomsItem], list[str]]:
        header_candidates = [
            word
            for word in full_words
            if 180 <= word.y0 <= 400
            and any(marker in _normalized_text(word.text) for marker in ("项号", "商品编号"))
        ]
        header_label = min(header_candidates, key=lambda word: (word.y0, word.x0)) if header_candidates else None
        table_y = header_label.y0 if header_label else default_table_y
        footer_markers = ("特殊关系", "支付特许权", "公式定价", "报关人员", "申报单位", "自报自缴")
        footer_y = min(
            (
                word.y0
                for word in full_words
                if word.y0 > table_y + 30 and any(marker in word.text for marker in footer_markers)
            ),
            default=TARGET_HEIGHT - 35,
        )
        bottom_y = max(table_y + 40, min(TARGET_HEIGHT - 25, table_y + 260, footer_y - 3))
        hs_words = self._ocr_crop(image, (55, table_y + 8, 138, bottom_y), "0123456789", enlarge=1.5)
        hs_words_alt = self._ocr_crop(image, (55, table_y + 8, 138, bottom_y), "0123456789", enlarge=1.5, threshold=185)
        code_entries = self._ocr_hs_entries(hs_words)
        alternate_entries = self._ocr_hs_entries(hs_words_alt)
        codes = [entry[1] for entry in code_entries]
        alternate_codes = [entry[1] for entry in alternate_entries]
        quantity_words = self._ocr_crop(image, (315, table_y, 425, bottom_y), "0123456789.件条千克公斤")
        price_words = self._ocr_crop(image, (420, table_y, 525, bottom_y), "0123456789.")
        product_words = self._ocr_crop(image, (100, table_y, 390, bottom_y))
        quantities = self._ocr_quantities(quantity_words)
        prices = self._ocr_prices(price_words)
        row_bounds = self._ocr_row_bounds(code_entries, table_y, bottom_y)
        warnings: list[str] = []
        if alternate_codes and alternate_codes != codes:
            warnings.append(f"两种OCR预处理得到的商品编号不一致：{codes} / {alternate_codes}")

        default_currency = self._detected_currency(" ".join(word.text for word in full_words))
        items: list[CustomsItem] = []
        assigned_quantities = 0
        assigned_prices = 0
        for index, ((code_y, code, code_confidence), (start_y, end_y)) in enumerate(zip(code_entries, row_bounds)):
            quantity_group = self._ocr_entries_in_row(quantities, start_y, end_y)
            price_group = self._ocr_entries_in_row(prices, start_y, end_y)
            assigned_quantities += len(quantity_group)
            assigned_prices += len(price_group)
            product_name, specification = self._ocr_product_text(product_words, start_y, end_y)
            origin_text = _region_text(full_words, 520, 620, start_y, end_y)
            destination_text = _region_text(full_words, 610, 700, start_y, end_y)
            source_text = _region_text(full_words, 680, 770, start_y, end_y)
            block_text = _region_text(full_words, 500, 835, start_y, end_y)
            row_text = _region_text(full_words, 400, 540, start_y, end_y)
            currency = self._detected_currency(row_text) or default_currency
            item_notes = ["扫描页OCR结果，商品名称和规格需人工复核"]
            if len(quantity_group) != 3:
                item_notes.append(f"数量及单位OCR得到{len(quantity_group)}组，预期3组")
            if len(price_group) != 2:
                item_notes.append(f"价格OCR得到{len(price_group)}个值，预期2个")
            if code_confidence < 0.8:
                item_notes.append(f"商品编号OCR置信度偏低：{code_confidence:.2f}")
            if any(entry[2] < 0.8 for entry in quantity_group + price_group):
                item_notes.append("数量或价格OCR置信度偏低")
            items.append(
                CustomsItem(
                    item_no=index + 1,
                    hs_code=code,
                    product_name=product_name,
                    specification=specification,
                    quantities=[quantity for _, quantity, _ in quantity_group[:3]],
                    unit_price=price_group[0][1] if price_group else None,
                    total_price=price_group[1][1] if len(price_group) > 1 else None,
                    currency=currency,
                    origin=origin_text or "中国 (CHN)",
                    final_destination=destination_text,
                    domestic_source=source_text.replace("照章征税", "").strip(),
                    exemption="照章征税 (1)" if "照章征税" in block_text else "",
                    review_notes=item_notes,
                )
            )
        unused_quantities = len(quantities) - assigned_quantities
        unused_prices = len(prices) - assigned_prices
        if unused_quantities or unused_prices:
            warnings.append(f"存在未匹配OCR字段：数量{unused_quantities}，价格{unused_prices}")
        if not items:
            warnings.append("OCR未能形成完整商品明细")
        return items, warnings

    def _ocr_header_numbers(self, image) -> tuple[int | None, Decimal | None, Decimal | None]:
        words = self._ocr_crop(image, (210, 160, 520, 220), "0123456789.")
        values: list[tuple[float, Decimal]] = []
        for word in words:
            if word.y0 < 190:
                continue
            match = re.fullmatch(r"\d+(?:\.\d+)?", word.text)
            value = _decimal(match.group(0)) if match else None
            if value is not None:
                values.append(((word.x0 + word.x1) / 2, value))
        assigned: list[Decimal | None] = []
        for target in (270.0, 317.0, 396.0):
            nearby = [entry for entry in values if abs(entry[0] - target) < 35]
            assigned.append(min(nearby, key=lambda entry: abs(entry[0] - target))[1] if nearby else None)
        packages_value = assigned[0]
        packages = (
            int(packages_value)
            if packages_value is not None and packages_value == packages_value.to_integral_value()
            else None
        )
        return packages, assigned[1], assigned[2]

    def _extract_with_ocr(self, document: fitz.Document, source_path: Path) -> ExtractionResult:
        selected_images: dict[int, object] = {}
        try:
            analyses: list[_OcrPageAnalysis] = []
            page_warnings: list[str] = []
            for page_index, page in enumerate(document):
                image = None
                try:
                    image = _extract_largest_page_image(document, page)
                    words = self._ocr_words(image)
                    analyses.append(self._analyze_ocr_page(page_index, words))
                except Exception as exc:
                    page_warnings.append(f"第{page_index + 1}页OCR失败：{exc}")
                finally:
                    if image is not None:
                        image.close()

            selected_pages, selection_warnings = self._select_ocr_pages(analyses)
            if not selected_pages:
                messages = ["逐页OCR后未找到可靠的报关单标题或商品表格"]
                messages.extend(page_warnings)
                return ExtractionResult(source_path.name, ExtractionStatus.OCR_REQUIRED, messages=messages)

            ocr_warnings = selection_warnings + page_warnings
            items: list[CustomsItem] = []
            for analysis in selected_pages:
                try:
                    image = _extract_largest_page_image(document, document[analysis.page_index])
                    selected_images[analysis.page_index] = image
                    code_words = [
                        word
                        for word in analysis.words
                        if 45 <= word.x0 < 150
                        and 180 <= word.y0 <= 450
                        and re.match(r"\s*\d{10,}", word.text)
                    ]
                    default_table_y = max(35.0, min(word.y0 for word in code_words) - 20) if code_words else 270.0
                    page_items, item_warnings = self._extract_ocr_items(
                        image,
                        analysis.words,
                        default_table_y=default_table_y,
                    )
                    for item in page_items:
                        item.item_no = len(items) + 1
                        items.append(item)
                    ocr_warnings.extend(
                        f"第{analysis.page_index + 1}页：{warning}" for warning in item_warnings
                    )
                except Exception as exc:
                    ocr_warnings.append(f"第{analysis.page_index + 1}页商品明细OCR失败：{exc}")

            header_page = max(
                selected_pages,
                key=lambda page: (
                    bool(page.customs_number),
                    page.header_hits,
                    page.has_title,
                    page.score,
                    -page.page_index,
                ),
            )
            all_text = "\n".join(page.text for page in selected_pages)
            declaration = _extract_header(
                source_path=source_path,
                pdf_pages=len(document),
                declaration_pages=[page.page_index + 1 for page in selected_pages],
                words=header_page.words,
                all_text=all_text,
                items=items,
            )

            packages: int | None = None
            gross_weight: Decimal | None = None
            net_weight: Decimal | None = None
            number_pages = sorted(
                selected_pages,
                key=lambda page: (page.page_index != header_page.page_index, page.page_index),
            )
            for page in number_pages:
                image = selected_images.get(page.page_index)
                if image is None:
                    continue
                page_packages, page_gross_weight, page_net_weight = self._ocr_header_numbers(image)
                packages = packages if packages is not None else page_packages
                gross_weight = gross_weight if gross_weight is not None else page_gross_weight
                net_weight = net_weight if net_weight is not None else page_net_weight
                if packages is not None and gross_weight is not None and net_weight is not None:
                    break
            declaration.packages = packages if packages is not None else declaration.packages
            declaration.gross_weight_kg = gross_weight if gross_weight is not None else declaration.gross_weight_kg
            declaration.net_weight_kg = net_weight if net_weight is not None else declaration.net_weight_kg
            declaration.review_notes.append("扫描版报关单使用EasyOCR识别，关键字段需复核")
            declaration.review_notes.extend(ocr_warnings)
            declaration.review_notes.extend(_validate_declaration(declaration))
            declaration.review_notes = list(dict.fromkeys(declaration.review_notes))
            return ExtractionResult(
                source_path.name,
                ExtractionStatus.REVIEW,
                declaration,
                declaration.review_notes.copy(),
            )
        except Exception as exc:
            return ExtractionResult(
                source_path.name,
                ExtractionStatus.OCR_REQUIRED,
                messages=[f"OCR识别失败：{exc}"],
            )
        finally:
            for image in selected_images.values():
                image.close()


def _xlsx_value(value):
    if isinstance(value, Decimal):
        return float(value)
    return value


def _set_widths(sheet, widths: Sequence[float]) -> None:
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width


def _wrapped_row_height(
    values: Sequence[tuple[object, int]],
    minimum: float,
    maximum: float,
) -> float:
    max_lines = 1
    for value, column_width in values:
        if value in (None, ""):
            continue
        line_count = 0
        for paragraph in str(value).splitlines() or [""]:
            display_width = sum(2 if east_asian_width(character) in "WFA" else 1 for character in paragraph)
            line_count += max(1, (display_width + column_width - 1) // column_width)
        max_lines = max(max_lines, line_count)
    return min(maximum, max(minimum, 7 + 13 * max_lines))


def _style_sheet(sheet, title: str, end_column: int, header_row: int) -> None:
    end_letter = get_column_letter(end_column)
    sheet.merge_cells(f"A1:{end_letter}1")
    sheet["A1"] = title
    sheet["A1"].fill = PatternFill("solid", fgColor=PRIMARY)
    sheet["A1"].font = Font(name="微软雅黑", size=16, bold=True, color=WHITE)
    sheet["A1"].alignment = Alignment(horizontal="left", vertical="center")
    sheet.row_dimensions[1].height = 30
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = f"A{header_row + 1}"


def _style_table_range(sheet, header_row: int, max_row: int, max_col: int, wrap_columns: set[int]) -> None:
    thin = Side(style="thin", color=GRID)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in sheet[header_row]:
        if cell.column > max_col:
            break
        cell.fill = PatternFill("solid", fgColor=PRIMARY)
        cell.font = Font(name="微软雅黑", size=9, bold=True, color=WHITE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    sheet.row_dimensions[header_row].height = 34
    for row in sheet.iter_rows(min_row=header_row + 1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.fill = PatternFill("solid", fgColor=WHITE if cell.row % 2 else LIGHT)
            cell.font = Font(name="微软雅黑", size=9, color=TEXT)
            cell.border = border
            cell.alignment = Alignment(
                horizontal="left" if isinstance(cell.value, str) else "right",
                vertical="center",
                wrap_text=cell.column in wrap_columns,
            )


def _add_table(sheet, name: str, ref: str) -> None:
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)


def export_excel(batch: ExtractionBatch, output_path: str | os.PathLike[str]) -> Path:
    declarations = batch.declarations
    if not declarations:
        raise ValueError("没有可导出的报关单；请查看提取错误信息")

    output = Path(output_path).expanduser().resolve()
    if output.suffix.lower() != ".xlsx":
        raise ValueError("输出文件必须使用 .xlsx 扩展名")
    output.parent.mkdir(parents=True, exist_ok=True)
    status_by_source = {result.source_pdf: result.status.value for result in batch.results}

    workbook = Workbook()
    workbook.properties.creator = "FinanceTool"
    workbook.properties.title = "海关出口报关单提取"

    summary = workbook.active
    summary.title = "报关单汇总"
    summary_headers = [
        "源PDF", "提取状态", "PDF总页数", "报关单页码", "预录入编号", "海关编号", "合同协议号",
        "境内发货人", "境外收货人", "出境关别", "出口日期", "申报日期", "运输方式", "运输工具名称及航次号",
        "提运单号", "监管方式", "征免性质", "贸易国(地区)", "运抵国(地区)", "指运港", "离境口岸",
        "包装种类", "件数", "毛重(kg)", "净重(kg)", "成交方式", "运费", "保费", "杂费", "商品项数",
        "币制", "申报总价", "备注", "复核备注",
    ]
    _style_sheet(summary, "海关出口报关单提取汇总", len(summary_headers), 3)
    summary["A2"] = f"报关单 {len(declarations)} 份    商品明细 {len(batch.items)} 行    需复核 {batch.review_count} 份"
    summary.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(summary_headers))
    summary["A2"].fill = PatternFill("solid", fgColor=SOFT)
    summary["A2"].font = Font(name="微软雅黑", size=9, color=TEXT)
    summary["A2"].alignment = Alignment(horizontal="left", vertical="center")
    for column, header in enumerate(summary_headers, start=1):
        summary.cell(3, column, header)

    for row_index, declaration in enumerate(declarations, start=4):
        totals = declaration.declared_totals
        currency = next(iter(totals), "") if len(totals) == 1 else " / ".join(totals)
        total = next(iter(totals.values()), None) if len(totals) == 1 else None
        review_text = "；".join(declaration.review_notes)
        values = [
            declaration.source_pdf,
            status_by_source.get(declaration.source_pdf, ""),
            declaration.pdf_pages,
            ",".join(str(page) for page in declaration.declaration_pages),
            declaration.pre_entry_number,
            declaration.customs_number,
            declaration.contract_number,
            declaration.domestic_consignor,
            declaration.foreign_consignee,
            declaration.exit_customs,
            declaration.export_date,
            declaration.declaration_date,
            declaration.transport_mode,
            declaration.transport_tool,
            declaration.bill_of_lading_no,
            declaration.supervision_mode,
            declaration.levy_nature,
            declaration.trade_country,
            declaration.destination_country,
            declaration.destination_port,
            declaration.exit_port,
            declaration.package_type,
            declaration.packages,
            declaration.gross_weight_kg,
            declaration.net_weight_kg,
            declaration.terms,
            declaration.freight,
            declaration.insurance,
            declaration.miscellaneous,
            len(declaration.items),
            currency,
            total,
            declaration.remarks,
            review_text,
        ]
        for column, value in enumerate(values, start=1):
            summary.cell(row_index, column, _xlsx_value(value))
        summary.row_dimensions[row_index].height = _wrapped_row_height(
            (
                (declaration.domestic_consignor, 25),
                (declaration.foreign_consignee, 25),
                (declaration.remarks, 36),
                (review_text, 48),
            ),
            minimum=32,
            maximum=180,
        )
        for column in (5, 6, 7):
            summary.cell(row_index, column).number_format = "@"
        for column in (11, 12):
            summary.cell(row_index, column).number_format = "yyyy-mm-dd"
        for column in (24, 25, 27, 28, 29, 32):
            summary.cell(row_index, column).number_format = "#,##0.00##"

    summary_max_row = 3 + len(declarations)
    _style_table_range(summary, 3, summary_max_row, len(summary_headers), {1, 4, 7, 8, 9, 10, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 33, 34})
    _add_table(summary, "CustomsSummaryTable", f"A3:{get_column_letter(len(summary_headers))}{summary_max_row}")
    summary.conditional_formatting.add(
        f"B4:B{summary_max_row}",
        FormulaRule(formula=['B4="需复核"'], fill=PatternFill("solid", fgColor=WARNING), font=Font(color=WARNING_TEXT, bold=True)),
    )
    _set_widths(summary, [18, 11, 9, 12, 22, 22, 16, 25, 25, 18, 12, 12, 17, 23, 22, 18, 18, 16, 16, 18, 20, 24, 8, 11, 11, 12, 10, 10, 10, 10, 10, 15, 36, 48])

    detail = workbook.create_sheet("商品明细")
    detail_headers = [
        "源PDF", "海关编号", "合同协议号", "申报日期", "项号", "商品编号", "商品名称", "规格型号",
        "数量1", "单位1", "数量2", "单位2", "数量3", "单位3", "单价", "总价", "币制",
        "原产国(地区)", "最终目的国(地区)", "境内货源地", "征免", "复核备注",
    ]
    _style_sheet(detail, "海关出口报关单商品明细", len(detail_headers), 3)
    detail["A2"] = "数量及单位1/2/3按PDF页面自上而下保留；OCR或源文件异常记录在复核备注。"
    detail.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(detail_headers))
    detail["A2"].fill = PatternFill("solid", fgColor=SOFT)
    detail["A2"].font = Font(name="微软雅黑", size=9, color=TEXT)
    for column, header in enumerate(detail_headers, start=1):
        detail.cell(3, column, header)

    detail_row = 4
    for declaration in declarations:
        for item in declaration.items:
            quantities: list[Quantity | None] = item.quantities[:3] + [None] * max(0, 3 - len(item.quantities))
            review_text = "；".join(item.review_notes)
            values = [
                declaration.source_pdf,
                declaration.customs_number,
                declaration.contract_number,
                declaration.declaration_date,
                item.item_no,
                item.hs_code,
                item.product_name,
                item.specification,
                quantities[0].value if quantities[0] else None,
                quantities[0].unit if quantities[0] else "",
                quantities[1].value if quantities[1] else None,
                quantities[1].unit if quantities[1] else "",
                quantities[2].value if quantities[2] else None,
                quantities[2].unit if quantities[2] else "",
                item.unit_price,
                item.total_price,
                item.currency,
                item.origin,
                item.final_destination,
                item.domestic_source,
                item.exemption,
                review_text,
            ]
            for column, value in enumerate(values, start=1):
                detail.cell(detail_row, column, _xlsx_value(value))
            detail.cell(detail_row, 2).number_format = "@"
            detail.cell(detail_row, 6).number_format = "@"
            detail.cell(detail_row, 4).number_format = "yyyy-mm-dd"
            for column in (9, 11, 13):
                detail.cell(detail_row, column).number_format = "#,##0.####"
            detail.cell(detail_row, 15).number_format = "0.0000"
            detail.cell(detail_row, 16).number_format = "#,##0.00##"
            detail.row_dimensions[detail_row].height = _wrapped_row_height(
                (
                    (item.product_name, 18),
                    (item.specification, 58),
                    (item.final_destination, 20),
                    (item.domestic_source, 19),
                    (review_text, 48),
                ),
                minimum=34,
                maximum=140,
            )
            detail_row += 1

    detail_max_row = detail_row - 1
    _style_table_range(detail, 3, detail_max_row, len(detail_headers), {1, 3, 6, 7, 8, 17, 18, 19, 20, 21, 22})
    if detail_max_row >= 4:
        _add_table(detail, "CustomsItemsTable", f"A3:{get_column_letter(len(detail_headers))}{detail_max_row}")
        detail.conditional_formatting.add(
            f"V4:V{detail_max_row}",
            FormulaRule(formula=["LEN(V4)>0"], fill=PatternFill("solid", fgColor=WARNING), font=Font(color=WARNING_TEXT)),
        )
    _set_widths(detail, [18, 22, 16, 12, 8, 14, 18, 58, 12, 9, 12, 9, 12, 9, 12, 14, 9, 18, 20, 19, 17, 48])

    log_sheet = workbook.create_sheet("提取结果")
    log_headers = ["源PDF", "提取状态", "海关编号", "商品项数", "信息"]
    _style_sheet(log_sheet, "提取结果与错误信息", len(log_headers), 3)
    log_sheet["A2"] = "失败或需要OCR的文件不会生成错误的报关单行；请按信息处理后重新运行。"
    log_sheet.merge_cells("A2:E2")
    log_sheet["A2"].fill = PatternFill("solid", fgColor=SOFT)
    log_sheet["A2"].font = Font(name="微软雅黑", size=9, color=TEXT)
    for column, header in enumerate(log_headers, start=1):
        log_sheet.cell(3, column, header)
    for row_index, result in enumerate(batch.results, start=4):
        message_text = "；".join(result.messages)
        values = [
            result.source_pdf,
            result.status.value,
            result.declaration.customs_number if result.declaration else "",
            len(result.declaration.items) if result.declaration else 0,
            message_text,
        ]
        for column, value in enumerate(values, start=1):
            log_sheet.cell(row_index, column, value)
        log_sheet.cell(row_index, 3).number_format = "@"
        log_sheet.row_dimensions[row_index].height = _wrapped_row_height(
            ((message_text, 90),),
            minimum=32,
            maximum=180,
        )
    log_max_row = 3 + len(batch.results)
    _style_table_range(log_sheet, 3, log_max_row, len(log_headers), {1, 2, 3, 5})
    _add_table(log_sheet, "ExtractionLogTable", f"A3:E{log_max_row}")
    _set_widths(log_sheet, [22, 12, 24, 12, 90])

    for sheet in workbook.worksheets:
        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.page_setup.orientation = "landscape"
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 0

    temporary_path: Path | None = None
    try:
        with tempfile.NamedTemporaryFile(
            prefix=f".{output.stem}-",
            suffix=".xlsx",
            dir=output.parent,
            delete=False,
        ) as temporary:
            temporary_path = Path(temporary.name)
        workbook.save(temporary_path)
        verification = load_workbook(temporary_path, read_only=False, data_only=False)
        if verification.sheetnames != ["报关单汇总", "商品明细", "提取结果"]:
            raise RuntimeError("导出的工作簿结构校验失败")
        if verification["报关单汇总"].max_row != summary_max_row:
            raise RuntimeError("报关单汇总行数校验失败")
        if verification["商品明细"].max_row != detail_max_row:
            raise RuntimeError("商品明细行数校验失败")
        verification.close()
        os.replace(temporary_path, output)
        temporary_path = None
    finally:
        if temporary_path and temporary_path.exists():
            temporary_path.unlink()
    return output


def _build_cli_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="从中国海关出口报关单PDF提取汇总和商品明细到Excel")
    parser.add_argument("inputs", nargs="+", help="PDF文件或包含PDF的目录")
    parser.add_argument("-o", "--output", default="报关单提取.xlsx", help="输出xlsx路径")
    parser.add_argument("--no-ocr", action="store_true", help="不启用EasyOCR；扫描页标记为需要OCR")
    parser.add_argument("--allow-review", action="store_true", help="存在需复核结果时仍返回退出码0")
    return parser


def main(argv: Sequence[str] | None = None) -> int:
    args = _build_cli_parser().parse_args(argv)
    extractor = CustomsDeclarationExtractor(enable_ocr=not args.no_ocr)
    files = extractor.expand_inputs(args.inputs)
    if not files:
        print("未找到PDF文件")
        return 2

    def report(index: int, total: int, result: ExtractionResult) -> None:
        print(f"[{index}/{total}] {result.source_pdf}: {result.status.value}")

    batch = extractor.extract_files(files, progress_callback=report)
    try:
        output = export_excel(batch, args.output)
    except Exception as exc:
        print(f"导出失败：{exc}")
        return 2
    print(f"已输出：{output}")
    print(
        f"报关单 {len(batch.declarations)} 份，商品 {len(batch.items)} 行，"
        f"需复核 {batch.review_count} 份，失败/需OCR {batch.failure_count} 份"
    )
    if batch.failure_count:
        return 2
    if batch.review_count and not args.allow_review:
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
