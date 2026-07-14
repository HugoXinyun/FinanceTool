from __future__ import annotations

import os
from datetime import date
from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import load_workbook

from customs_declaration_extractor import (
    CustomsDeclaration,
    CustomsDeclarationExtractor,
    CustomsItem,
    ExtractionBatch,
    ExtractionResult,
    ExtractionStatus,
    PdfWord,
    Quantity,
    _date,
    _decimal,
    _validate_declaration,
    export_excel,
)


ROOT = Path(__file__).resolve().parents[1]
SAMPLE_DIR = ROOT / "docs" / "6月报关"


def require_samples() -> Path:
    if not SAMPLE_DIR.is_dir():
        pytest.skip("docs/6月报关 sample directory is not available")
    return SAMPLE_DIR


def make_item(**overrides) -> CustomsItem:
    values = {
        "item_no": 1,
        "hs_code": "6204690000",
        "product_name": "女士梭织短裤",
        "specification": "3|0|梭织|短裤|女式",
        "quantities": [Quantity(Decimal("2"), "条"), Quantity(Decimal("1"), "千克")],
        "unit_price": Decimal("10"),
        "total_price": Decimal("20"),
        "currency": "美元",
        "origin": "中国 (CHN)",
        "final_destination": "英国 (GBR)",
        "domestic_source": "(44019) 广州其他",
        "exemption": "照章征税 (1)",
    }
    values.update(overrides)
    return CustomsItem(**values)


def make_declaration(*, items: list[CustomsItem] | None = None) -> CustomsDeclaration:
    return CustomsDeclaration(
        source_pdf="source.pdf",
        pdf_pages=1,
        declaration_pages=[1],
        pre_entry_number="531720260000257302",
        customs_number="531720260000257302",
        contract_number="E3342049-138",
        domestic_consignor="广州伊珂贝科技有限公司",
        foreign_consignee="A.M. LONDON FASHION LTD",
        exit_customs="(0101) 京机场关",
        export_date=date(2026, 6, 26),
        declaration_date=date(2026, 6, 26),
        transport_mode="(5) 航空运输",
        transport_tool="",
        bill_of_lading_no="",
        supervision_mode="(0110) 一般贸易",
        levy_nature="(101) 一般征税",
        trade_country="(GBR) 英国",
        destination_country="(GBR) 英国",
        destination_port="(GBR000) 英国",
        exit_port="(110101) 首都国际机场",
        package_type="(22) 纸制或纤维板制盒/箱",
        packages=1,
        gross_weight_kg=Decimal("2"),
        net_weight_kg=Decimal("1"),
        terms="(3) FOB",
        freight=None,
        insurance=None,
        miscellaneous=None,
        remarks="",
        items=[make_item()] if items is None else items,
    )


def make_ocr_words(*texts: str) -> list[PdfWord]:
    return [
        PdfWord(10, index * 10, 300, index * 10 + 8, text)
        for index, text in enumerate(texts)
    ]


def test_scalar_parsers():
    assert _decimal("1,509.75") == Decimal("1509.75")
    assert _decimal(16) == Decimal("16")
    assert _decimal("") is None
    assert _date("20260626") == date(2026, 6, 26)
    assert _date("2026-06-26") == date(2026, 6, 26)
    assert _date("not-a-date") is None


def test_expand_inputs_keeps_first_pdf_and_ignores_other_files(tmp_path):
    first = tmp_path / "A.PDF"
    second = tmp_path / "b.pdf"
    other = tmp_path / "notes.txt"
    first.touch()
    second.touch()
    other.touch()

    files = CustomsDeclarationExtractor.expand_inputs([tmp_path, first])

    assert files == [first.resolve(), second.resolve()]


@pytest.mark.parametrize(
    ("item_overrides", "expected_note"),
    [
        ({"unit_price": None}, "第1项未识别单价"),
        ({"total_price": None}, "第1项未识别总价"),
        ({"currency": ""}, "第1项未识别币制"),
    ],
)
def test_missing_price_or_currency_requires_review(item_overrides, expected_note):
    declaration = make_declaration(items=[make_item(**item_overrides)])

    notes = _validate_declaration(declaration)

    assert expected_note in notes


def test_missing_net_weight_requires_review():
    declaration = make_declaration()
    declaration.net_weight_kg = None

    notes = _validate_declaration(declaration)

    assert "未识别报关单净重" in notes


def test_ocr_page_selection_finds_declaration_on_second_page():
    customs_number = "531720260000257302"
    analyses = [
        CustomsDeclarationExtractor._analyze_ocr_page(
            0,
            make_ocr_words("放行通知", f"海关编号：{customs_number}"),
        ),
        CustomsDeclarationExtractor._analyze_ocr_page(
            1,
            make_ocr_words(
                "中华人民共和国海关出口货物报关单",
                f"海关编号：{customs_number}",
                "项号",
                "商品编号",
                "6204690000",
            ),
        ),
    ]

    selected, warnings = CustomsDeclarationExtractor._select_ocr_pages(analyses)

    assert [page.page_index for page in selected] == [1]
    assert warnings == []


def test_ocr_page_selection_merges_adjacent_item_continuation_page():
    customs_number = "531720260000257302"
    analyses = [
        CustomsDeclarationExtractor._analyze_ocr_page(
            0,
            make_ocr_words(
                "中华人民共和国海关出口货物报关单",
                f"海关编号：{customs_number}",
                "项号",
                "商品编号",
                "6204690000",
            ),
        ),
        CustomsDeclarationExtractor._analyze_ocr_page(
            1,
            make_ocr_words(
                f"海关编号：{customs_number}",
                "项号",
                "商品编号",
                "6211499000",
            ),
        ),
    ]

    assert analyses[1].has_item_table

    selected, warnings = CustomsDeclarationExtractor._select_ocr_pages(analyses)

    assert [page.page_index for page in selected] == [0, 1]
    assert warnings == []


def test_ocr_page_selection_excludes_release_page_without_item_identity():
    customs_number = "531720260000257302"
    analyses = [
        CustomsDeclarationExtractor._analyze_ocr_page(
            0,
            make_ocr_words(
                "中华人民共和国海关出口货物报关单",
                f"海关编号：{customs_number}",
                "项号",
                "商品编号",
                "6204690000",
            ),
        ),
        CustomsDeclarationExtractor._analyze_ocr_page(
            1,
            make_ocr_words(
                "放行通知",
                f"海关编号：{customs_number}",
                "数量及单位",
                "单价",
                "总价",
            ),
        ),
    ]

    assert not analyses[1].has_item_identity
    assert not analyses[1].has_item_table

    selected, warnings = CustomsDeclarationExtractor._select_ocr_pages(analyses)

    assert [page.page_index for page in selected] == [0]
    assert warnings == []


def test_export_rejects_non_xlsx_without_overwriting_input(tmp_path):
    input_pdf = tmp_path / "source.pdf"
    original_contents = b"original PDF contents"
    input_pdf.write_bytes(original_contents)
    declaration = make_declaration()
    batch = ExtractionBatch(
        [ExtractionResult(declaration.source_pdf, ExtractionStatus.SUCCESS, declaration)]
    )

    with pytest.raises(ValueError, match=r"\.xlsx"):
        export_excel(batch, input_pdf)

    assert input_pdf.read_bytes() == original_contents


def test_export_allows_empty_item_details_and_keeps_extraction_log(tmp_path):
    declaration = make_declaration(items=[])
    message = "未识别商品明细"
    declaration.review_notes.append(message)
    batch = ExtractionBatch(
        [
            ExtractionResult(
                declaration.source_pdf,
                ExtractionStatus.REVIEW,
                declaration,
                messages=[message],
            )
        ]
    )

    output = export_excel(batch, tmp_path / "empty-items.xlsx")

    workbook = load_workbook(output, data_only=False)
    summary = workbook["报关单汇总"]
    detail = workbook["商品明细"]
    log = workbook["提取结果"]
    assert summary["AD4"].value == 0
    assert detail.max_row == 3
    assert len(detail.tables) == 0
    assert log["A4"].value == "source.pdf"
    assert log["B4"].value == ExtractionStatus.REVIEW.value
    assert log["D4"].value == 0
    assert log["E4"].value == message
    assert len(log.tables) == 1
    workbook.close()


def test_export_expands_rows_for_long_review_messages(tmp_path):
    message = "扫描件字段需要人工复核；" * 30
    declaration = make_declaration(items=[make_item(review_notes=[message])])
    declaration.review_notes.append(message)
    batch = ExtractionBatch(
        [
            ExtractionResult(
                declaration.source_pdf,
                ExtractionStatus.REVIEW,
                declaration,
                messages=[message],
            )
        ]
    )

    output = export_excel(batch, tmp_path / "long-review.xlsx")

    workbook = load_workbook(output, data_only=False)
    assert workbook["报关单汇总"].row_dimensions[4].height > 46
    assert workbook["商品明细"].row_dimensions[4].height > 56
    assert workbook["提取结果"].row_dimensions[4].height > 36
    workbook.close()


@pytest.mark.integration
def test_extract_text_pdf_batch_matches_known_totals():
    sample_dir = require_samples()
    files = [sample_dir / f"E3342049-{suffix}.pdf" for suffix in range(132, 138)]
    batch = CustomsDeclarationExtractor(enable_ocr=False).extract_files(files)

    assert len(batch.declarations) == 6
    assert len(batch.items) == 27
    expected = {
        "E3342049-132": (1, Decimal("225.5"), Decimal("5488.25")),
        "E3342049-133": (12, Decimal("766.0"), Decimal("20754.21")),
        "E3342049-134": (1, Decimal("2240.4"), Decimal("29043.36")),
        "E3342049-135": (2, Decimal("235.2"), Decimal("8863.61")),
        "E3342049-136": (3, Decimal("380.9"), Decimal("6057.6")),
        "E3342049-137": (8, Decimal("4113.3"), Decimal("67937.7")),
    }
    for declaration in batch.declarations:
        item_count, weight, amount = expected[declaration.contract_number]
        assert len(declaration.items) == item_count
        assert declaration.total_item_weight_kg == weight
        assert declaration.declared_totals["美元"] == amount
        assert declaration.net_weight_kg == weight
        assert isinstance(declaration.customs_number, str)
        assert len(declaration.customs_number) == 18
        assert all(isinstance(item.hs_code, str) and len(item.hs_code) == 10 for item in declaration.items)

    declaration_134 = next(item for item in batch.declarations if item.contract_number == "E3342049-134")
    assert declaration_134.items[0].quantities[2].unit == "千克"
    assert declaration_134.items[0].quantities[2].value == Decimal("2240.4")
    assert all(result.status not in {ExtractionStatus.FAILED, ExtractionStatus.OCR_REQUIRED} for result in batch.results)


@pytest.mark.integration
def test_scanned_pdf_requires_ocr_when_disabled():
    sample_dir = require_samples()
    result = CustomsDeclarationExtractor(enable_ocr=False).extract_pdf(sample_dir / "E3342049-138.pdf")

    assert result.status == ExtractionStatus.OCR_REQUIRED
    assert result.declaration is None
    assert "OCR" in " ".join(result.messages)


@pytest.mark.integration
def test_export_excel_preserves_codes_and_types(tmp_path):
    sample_dir = require_samples()
    batch = CustomsDeclarationExtractor(enable_ocr=False).extract_files([sample_dir / "E3342049-132.pdf"])
    output = export_excel(batch, tmp_path / "customs.xlsx")

    workbook = load_workbook(output, data_only=False)
    assert workbook.sheetnames == ["报关单汇总", "商品明细", "提取结果"]
    summary = workbook["报关单汇总"]
    detail = workbook["商品明细"]
    assert summary["F4"].value == "516620260000816924"
    assert isinstance(summary["F4"].value, str)
    assert summary["L4"].value.date() == date(2026, 6, 2)
    assert detail["F4"].value == "6110300090"
    assert isinstance(detail["F4"].value, str)
    assert detail["P4"].value == pytest.approx(5488.25)
    assert len(summary.tables) == 1
    assert len(detail.tables) == 1
    workbook.close()


@pytest.mark.integration
@pytest.mark.ocr
@pytest.mark.skipif(os.environ.get("RUN_OCR_TESTS") != "1", reason="set RUN_OCR_TESTS=1 to run EasyOCR")
def test_scanned_pdf_ocr_extracts_verified_numeric_fields():
    sample_dir = require_samples()
    result = CustomsDeclarationExtractor(enable_ocr=True).extract_pdf(sample_dir / "E3342049-138.pdf")

    assert result.status == ExtractionStatus.REVIEW
    declaration = result.declaration
    assert declaration is not None
    assert declaration.customs_number == "531720260000257302"
    assert declaration.packages == 16
    assert declaration.gross_weight_kg == Decimal("155")
    assert declaration.net_weight_kg == Decimal("139")
    assert [item.hs_code for item in declaration.items] == ["6204690000", "6211499000"]
    assert [
        (
            item.hs_code,
            [(quantity.value, quantity.unit) for quantity in item.quantities],
            item.unit_price,
            item.total_price,
        )
        for item in declaration.items
    ] == [
        (
            "6204690000",
            [(Decimal("305"), "条"), (Decimal("56.4"), "千克"), (Decimal("305"), "条")],
            Decimal("4.95"),
            Decimal("1509.75"),
        ),
        (
            "6211499000",
            [(Decimal("539"), "件"), (Decimal("82.6"), "千克"), (Decimal("539"), "件")],
            Decimal("7.5"),
            Decimal("4042.5"),
        ),
    ]
    assert declaration.total_item_weight_kg == Decimal("139.0")
    assert declaration.declared_totals["美元"] == Decimal("5552.25")
