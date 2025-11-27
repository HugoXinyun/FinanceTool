# excel_merge_tab.py

from PyQt5.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QPushButton, QListWidget,
    QFileDialog, QMessageBox, QListWidgetItem
)
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QDragEnterEvent
from openpyxl import load_workbook, Workbook
import os
import xlrd  # 用于读取 .xls 文件
from datetime import datetime
import subprocess
import time

# 尝试导入pyautogui，如果失败则标记为不可用
try:
    import pyautogui
    PYAUTOGUI_AVAILABLE = True
except ImportError:
    PYAUTOGUI_AVAILABLE = False
    pyautogui = None


class FileListWidget(QListWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setAcceptDrops(True)

    def dragEnterEvent(self, event: QDragEnterEvent):
        if event.mimeData().hasUrls():
            for url in event.mimeData().urls():
                file_path = url.toLocalFile()
                if file_path.endswith(('.xlsx', '.xls')):
                    modification_time = datetime.fromtimestamp(os.path.getmtime(file_path)).strftime('%Y-%m-%d %H:%M:%S')
                    file_name = os.path.basename(file_path)
                    item_text = f"{file_name} (修改日期: {modification_time})"
                    item = QListWidgetItem(item_text)
                    item.setData(Qt.UserRole, file_path)
                    self.addItem(item)
            event.acceptProposedAction()
        else:
            event.ignore()


class ExcelMergeTab(QWidget):
    def __init__(self):
        super().__init__()
        self.init_ui()

    def init_ui(self):
        layout = QHBoxLayout()

        left_layout = QVBoxLayout()
        right_layout = QVBoxLayout()

        self.file_list = FileListWidget()
        self.file_list.setSelectionMode(QListWidget.MultiSelection)
        left_layout.addWidget(self.file_list)

        self.add_button = QPushButton("添加Excel")
        self.add_button.clicked.connect(self.add_excel)
        right_layout.addWidget(self.add_button)

        self.delete_button = QPushButton("删除选择")
        self.delete_button.clicked.connect(self.delete_excel)
        right_layout.addWidget(self.delete_button)

        self.clear_button = QPushButton("清除列表")
        self.clear_button.clicked.connect(self.clear_list)
        right_layout.addWidget(self.clear_button)

        self.remove_duplicates_button = QPushButton("清除重复项")
        self.remove_duplicates_button.clicked.connect(self.remove_duplicates)
        right_layout.addWidget(self.remove_duplicates_button)

        self.top_button = QPushButton("置顶")
        self.top_button.clicked.connect(self.move_to_top)
        right_layout.addWidget(self.top_button)

        self.up_button = QPushButton("上移")
        self.up_button.clicked.connect(self.move_up)
        right_layout.addWidget(self.up_button)

        self.down_button = QPushButton("下移")
        self.down_button.clicked.connect(self.move_down)
        right_layout.addWidget(self.down_button)

        self.bottom_button = QPushButton("置底")
        self.bottom_button.clicked.connect(self.move_to_bottom)
        right_layout.addWidget(self.bottom_button)

        self.merge_button = QPushButton("合并(9行后)")
        self.merge_button.clicked.connect(self.merge_files)
        right_layout.addWidget(self.merge_button)

        self.print_button = QPushButton("打印列表文件")
        self.print_button.clicked.connect(self.print_files)
        right_layout.addWidget(self.print_button)
        layout.addLayout(left_layout)
        layout.addLayout(right_layout)
        self.setLayout(layout)

    def add_excel(self):
        files, _ = QFileDialog.getOpenFileNames(self, "选择Excel文件", "", "Excel Files (*.xlsx *.xls)")
        if files:
            for file in files:
                modification_time = datetime.fromtimestamp(os.path.getmtime(file)).strftime('%Y-%m-%d %H:%M:%S')
                file_name = os.path.basename(file)
                item_text = f"{file_name} (修改日期: {modification_time})"
                item = QListWidgetItem(item_text)
                item.setData(Qt.UserRole, file)
                self.file_list.addItem(item)

    def delete_excel(self):
        for item in self.file_list.selectedItems():
            self.file_list.takeItem(self.file_list.row(item))

    def clear_list(self):
        self.file_list.clear()

    def remove_duplicates(self):
        excel_files = [self.file_list.item(i).data(Qt.UserRole) for i in range(self.file_list.count())]
        unique_files = []
        seen_files = set()
        for file in reversed(excel_files):
            if file not in seen_files:
                unique_files.append(file)
                seen_files.add(file)
        unique_files.reverse()

        self.file_list.clear()
        for file in unique_files:
            item_text = f"{os.path.basename(file)} (修改日期: {datetime.fromtimestamp(os.path.getmtime(file)).strftime('%Y-%m-%d %H:%M:%S')})"
            item = QListWidgetItem(item_text)
            item.setData(Qt.UserRole, file)
            self.file_list.addItem(item)

    def move_to_top(self):
        current_row = self.file_list.currentRow()
        if current_row > 0:
            item = self.file_list.takeItem(current_row)
            self.file_list.insertItem(0, item)
            self.file_list.setCurrentRow(0)

    def move_up(self):
        current_row = self.file_list.currentRow()
        if current_row > 0:
            item = self.file_list.takeItem(current_row)
            self.file_list.insertItem(current_row - 1, item)
            self.file_list.setCurrentRow(current_row - 1)

    def move_down(self):
        current_row = self.file_list.currentRow()
        if current_row < self.file_list.count() - 1:
            item = self.file_list.takeItem(current_row)
            self.file_list.insertItem(current_row + 1, item)
            self.file_list.setCurrentRow(current_row + 1)

    def move_to_bottom(self):
        current_row = self.file_list.currentRow()
        if current_row < self.file_list.count() - 1:
            item = self.file_list.takeItem(current_row)
            self.file_list.addItem(item)
            self.file_list.setCurrentRow(self.file_list.count() - 1)

    def merge_files(self):
        output_path, _ = QFileDialog.getSaveFileName(self, "保存合并后的Excel", "", "Excel Files (*.xlsx)")
        if not output_path:
            return

        excel_files = [self.file_list.item(i).data(Qt.UserRole) for i in range(self.file_list.count())]
        if not excel_files:
            QMessageBox.warning(self, "警告", "请先添加Excel文件")
            return
        # 去除重复文件
        unique_files = []
        seen_files = set()
        for file in reversed(excel_files):
            if file not in seen_files:
                unique_files.append(file)
                seen_files.add(file)
        unique_files.reverse()

        try:
            merged_wb = Workbook()
            default_sheet = merged_wb.active
            default_sheet.title = "MergedData"

            # 获取第一个文件的第一个工作表名称作为目标工作表名称
            if unique_files:
                first_file = unique_files[0]
                _, ext_first = os.path.splitext(first_file)
                if ext_first.lower() == '.xlsx':
                    wb_first = load_workbook(first_file)
                    first_sheet_name = wb_first.sheetnames[0]
                    ws_first = wb_first[first_sheet_name]
                    header_rows = list(ws_first.iter_rows(min_row=1, max_row=8, values_only=True))
                elif ext_first.lower() == '.xls':
                    wb_first = xlrd.open_workbook(first_file)
                    first_sheet_name = wb_first.sheet_names()[0]
                    ws_first = wb_first.sheet_by_name(first_sheet_name)
                    header_rows = [ws_first.row_values(row_idx, 0, ws_first.ncols) for row_idx in range(0, 8)]
                else:
                    first_sheet_name = "Sheet1"
                    header_rows = []

            merged_sheet = merged_wb.create_sheet(title=first_sheet_name)
            merged_wb.remove(merged_wb["MergedData"])  # 删除默认创建的工作表

            # 插入第一个文件的前8行
            for row in header_rows:
                merged_sheet.append(row)

            for file in unique_files:
                _, ext = os.path.splitext(file)
                if ext.lower() == '.xlsx':
                    wb = load_workbook(file)
                    for sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]

                        # 从第9行开始读取，直到最后一个非空行
                        for row in ws.iter_rows(min_row=9, values_only=True):
                            if any(cell is not None and cell != '' for cell in row):
                                merged_sheet.append(row)
                elif ext.lower() == '.xls':
                    wb = xlrd.open_workbook(file)
                    for sheet_name in wb.sheet_names():
                        ws = wb.sheet_by_name(sheet_name)

                        # 从第9行开始读取，直到最后一个非空行
                        for row_idx in range(8, ws.nrows):
                            row = ws.row_values(row_idx)
                            if any(cell != "" for cell in row):
                                merged_sheet.append(row)

            merged_wb.save(output_path)
            QMessageBox.information(self, "成功", f"Excel文件已成功合并到:\n{output_path}")

        except Exception as e:
            QMessageBox.critical(self, "错误", f"合并Excel失败:\n{str(e)}")

    def print_files(self):
        try:
            current_time = datetime.now().strftime("%y%m%d-%H%M%S")
            output_path = os.path.abspath(f"{current_time}.xlsx")

            excel_files = [self.file_list.item(i).data(Qt.UserRole) for i in range(self.file_list.count())]
            if not excel_files:
                QMessageBox.warning(self, "警告", "请先添加Excel文件")
                return

            unique_files = []
            seen_files = set()
            for file in reversed(excel_files):
                if file not in seen_files:
                    unique_files.append(file)
                    seen_files.add(file)
            unique_files.reverse()

            # 合并逻辑
            merged_wb = Workbook()
            default_sheet = merged_wb.active
            merged_wb.remove(default_sheet)

            for file in unique_files:
                _, ext = os.path.splitext(file)
                if ext.lower() == '.xlsx':
                    wb = load_workbook(file)
                    for sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                        new_ws = merged_wb.create_sheet(f"{os.path.splitext(os.path.basename(file))[0]}_{sheet_name}")
                        for row in ws.iter_rows(values_only=True):
                            new_ws.append(row)
                elif ext.lower() == '.xls':
                    wb = xlrd.open_workbook(file)
                    for sheet_name in wb.sheet_names():
                        ws = wb.sheet_by_name(sheet_name)
                        new_ws = merged_wb.create_sheet(f"{os.path.splitext(os.path.basename(file))[0]}_{sheet_name}")
                        for row_idx in range(ws.nrows):
                            row = ws.row_values(row_idx)
                            new_ws.append(row)

            merged_wb.save(output_path)

            if not os.path.exists(output_path):
                QMessageBox.critical(self, "错误", "生成的Excel文件不存在")
                return

            # 使用默认程序打开 Excel 文件
            subprocess.Popen(['start', 'excel', output_path], shell=True)
            time.sleep(2)  # 等待Excel加载
            
            # 检查pyautogui是否可用
            if PYAUTOGUI_AVAILABLE and pyautogui:
                pyautogui.hotkey('ctrl', 'p')  # 触发打印
                QMessageBox.information(self, "成功", "打印任务已启动")
            else:
                QMessageBox.information(self, "提示", f"Excel文件已生成并打开:\n{output_path}\n请手动按 Ctrl+P 打印")

        except Exception as e:
            QMessageBox.critical(self, "错误", f"打印失败:\n{str(e)}")
